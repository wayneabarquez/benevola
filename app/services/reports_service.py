from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Style, Font, Alignment
from openpyxl.styles.borders import Border, Side
from app.utils.date_utils import convert_24_hour_to_12_hour, convert_date_to_human_readable
from app import app
from app.services import section_service
import logging
import datetime
from operator import itemgetter

log = logging.getLogger(__name__)

LOT_LIST_FOLDER = app.config['LOT_LIST_REPORTS_FOLDER']
CEMETERY_NAME = app.config['CEMETERY_NAME']
CEMETERY_LOCATION = app.config['CEMETERY_LOCATION']
CREMATION_LIST_FOLDER = app.config['CREMATION_LIST_REPORTS_FOLDER']

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

total_border = Border(bottom=Side(style='double'))


def print_row_spacer(ctr, ws):
    for i in range(ctr):
        ws.append([])


def create_bordered_cell(val, ws, is_bold=False):
    c = Cell(ws, value=val)
    c.font = Font(size=11, bold=is_bold)
    c.border = thin_border
    return c


def underline_border_cell(val, ws):
    underline_border = Border(bottom=Side(style='thin'))

    c = Cell(ws, value=val)
    c.font = Font(size=11, bold=True)
    c.border = underline_border
    return c


def generate_file_name(report_type):
    return report_type + datetime.datetime.now().strftime('%Y-%m-%d')


def generate_lot_list_file_name():
    return generate_file_name('LOT_LIST_')


def generate_sales_file_name():
    return generate_file_name('SALES_')


def print_worksheet_header(ws):
    # Cemetery Name
    ws.merge_cells('D2:H2')
    ws['D2'].style = Style(font=Font(size=14, bold=True), alignment=Alignment(horizontal="center"))
    ws['D2'] = CEMETERY_NAME

    # Cemetery Location
    ws.merge_cells('D3:H3')
    ws['D3'].style = Style(font=Font(bold=True), alignment=Alignment(horizontal="center"))
    ws['D3'] = CEMETERY_LOCATION

    print_row_spacer(2, ws)


# LOT LIST REPORT FUNCTIONS

def print_lot_list_sections_data(ws):
    sections = section_service.get_sections()

    for section in sections:
        lots = section.get_lots()
        sold_lots = filter(lambda lot: lot['status'] == 'sold' or lot['status'] == 'occupied', lots)
        unsold_lots = filter(lambda lot: lot['status'] != 'sold' and lot['status'] != 'occupied', lots)

        # SECTION NAME ROW
        section_name = 'SECTION ' + section.name
        section_name_cell = Cell(ws, column='C', value=section_name)
        section_name_cell.font = Font(size=14, bold=True)
        section_name_cell.border = thin_border
        ws.append([None, None, section_name_cell])

        # NO OF BLOCKS ROW
        no_block_cell = create_bordered_cell('NO. OF BLOCKS', ws, True)
        no_block_value_cell = create_bordered_cell(len(section.blocks), ws)

        ws.append([None, None, no_block_cell, no_block_value_cell])

        # NO OF LOTS ROW
        no_lots_cell = create_bordered_cell('NO. OF LOTS', ws, True)
        no_lots_value_cell = create_bordered_cell(len(lots), ws)
        ws.append([None, None, no_lots_cell, no_lots_value_cell])

        # NO OF SOLD LOTS ROW
        no_sold_lots_cell = create_bordered_cell('SOLD LOTS', ws, True)
        no_sold_lots_value_cell = create_bordered_cell(len(sold_lots), ws)

        ws.append([None, None, no_sold_lots_cell, no_sold_lots_value_cell])

        # NO OF UNSOLD LOTS ROW
        no_unsold_lots_cell = create_bordered_cell('UNSOLD LOTS', ws, True)
        no_unsold_lots_value_cell = create_bordered_cell(len(unsold_lots), ws)

        ws.append([None, None, no_unsold_lots_cell, no_unsold_lots_value_cell])

        ws.append([])
        # Table Header
        table_headers = ['BLOCK', 'LOT NO.', 'DIMENSION', 'AREA', 'PRICE/SM', 'AMOUNT', 'REMARKS', 'OWNER',
                         'DATE PURCHASED']
        table_header_cells = []
        for h in table_headers:
            c = create_bordered_cell(h, ws, True)
            table_header_cells.append(c)
        r = [''] + table_header_cells
        ws.append(r)

        # TABLE BODY - LOTS
        # sorted_blocks = sorted(section.blocks, key=lambda block: block.name)
        for b in section.blocks:
            sorted_lots = sorted(b.lots, key=lambda lot: lot.name)
            for l in sorted_lots:
                client_name = l.client.get_full_name() if l.client is not None else ''
                price_formatted = '{:20,.2f}'.format(l.price_per_sq_mtr)
                amount = '{:20,.2f}'.format(l.lot_area * l.price_per_sq_mtr)
                row_val = [b.name, l.name, l.dimension, l.lot_area, price_formatted, amount, l.remarks, client_name,
                           l.date_purchased]
                row_cells = []
                for rv in row_val:
                    row_cells.append(create_bordered_cell(rv, ws))

                ws.append([''] + row_cells)
            # Total of Lots per block
            ws.append([len(b.lots)])

        print_row_spacer(2, ws)


def generate_lotlist_report():
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 15

    print_worksheet_header(ws)

    print_lot_list_sections_data(ws)

    filename = generate_lot_list_file_name() + ".xlsx"
    file_dict = {'name': filename, 'dir': LOT_LIST_FOLDER + filename}
    # Save the file
    wb.save(file_dict['dir'])

    return file_dict


# SALES REPORT FUNCTIONS

def print_sales_content(collection, ws):
    # Table Header
    table_headers = ['Date', 'O.R. #', 'Name', 'Amount', 'Remarks']
    table_header_cells = []

    for h in table_headers:
        c = underline_border_cell(h, ws)
        table_header_cells.append(c)
    r = [''] + table_header_cells
    ws.append(r)

    sales_total = 0
    for item in collection:
        if item.label == 'Cemetery Lot':
            amount = item.lot_area * item.price_per_sq_mtr
        # elif item.label == 'Cremation': # todo no amount for cremation yet
        #     amount = 0
        elif item.label == 'Columbary':
            amount = item.price if item.price is not None else 0

        amount_formatted = 'P {:20,.2f}'.format(amount)
        amount_formatted_cell = Cell(ws, value=amount_formatted)
        amount_formatted_cell.style = Style(alignment=Alignment(horizontal='right'))
        client_name = item.client.get_full_name() if item.client is not None else ''

        sales_total += amount
        ws.append(['', item.date_purchased, item.or_no, client_name, amount_formatted_cell, item.label])

    # Sales Total
    total_label_cell = Cell(ws, value='TOTAL')
    total_label_cell.font = Font(size=12, color='FFFF0000')

    total_cell = Cell(ws, value='P {:20,.2f}'.format(sales_total))
    total_cell.font = Font(size=12, color='FFFF0000')
    total_cell.border = total_border
    total_cell.alignment = Alignment(horizontal='right')

    ws.append(['', '', '', total_label_cell, total_cell])


def print_cremation_list_content(collection, ws):
    # Table Header
    table_headers = ['DATE', 'NAME OF DECEASED CREMATED', 'SEX', 'AGE', 'TIME STARTED', 'TIME FINISHED', 'GAS (liters)']
    table_header_cells = []

    for h in table_headers:
        c = underline_border_cell(h, ws)
        table_header_cells.append(c)
    r = [''] + table_header_cells
    ws.append(r)

    for item in collection:
        client = item.funeral_homes.name if item.funeral_homes is not None else item.deceased.full_name.upper()
        ws.append(['', convert_date_to_human_readable(item.date_cremated), client,
                   item.deceased.gender[0].upper(), item.deceased.age, convert_24_hour_to_12_hour(item.time_started),
                   convert_24_hour_to_12_hour(item.time_finished), item.gas_consumed])


def generate_sales_report(start_date, end_date):
    from app.services import lot_service
    collection = lot_service.get_sales_data_by_date(start_date, end_date)

    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    print_worksheet_header(ws)

    # print_sales_content(lots, ws)
    print_sales_content(collection, ws)

    filename = generate_sales_file_name() + ".xlsx"
    file_dict = {'name': filename, 'dir': LOT_LIST_FOLDER + filename}
    # Save the file
    wb.save(file_dict['dir'])

    return file_dict


def generate_cremation_list_report(start_date, end_date):
    from app.crematorium.services import get_cremation_list_by_date
    collection = get_cremation_list_by_date(start_date, end_date)

    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    print_worksheet_header(ws)

    # print_sales_content(lots, ws)
    print_cremation_list_content(collection, ws)

    filename = generate_file_name('CREMATION_LIST_') + ".xlsx"
    file_dict = {'name': filename, 'dir': CREMATION_LIST_FOLDER + filename}
    # Save the file
    wb.save(file_dict['dir'])

    return file_dict
